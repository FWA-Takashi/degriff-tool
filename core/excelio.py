# -*- coding: utf-8 -*-
"""Read uploaded spreadsheets and build the CREATION / DISPATCH / Benchmark
output workbooks. Ported 1:1 from the old SheetJS JavaScript so the column
order and number formats match the gold templates exactly.
"""
import csv as _csv
import io

from openpyxl import Workbook, load_workbook

from .rules import EUR_FMT, EUR0_FMT, num, to_rayon


# ---------------------------------------------------------------- reading
def read_tabular(file_bytes, filename):
    """Return an AOA (list of rows, each a list of cell values; blanks -> "")."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", "ignore")
        return [list(r) for r in _csv.reader(io.StringIO(text))]
    # xlsx / xlsm
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    aoa = []
    for row in ws.iter_rows(values_only=True):
        aoa.append(["" if c is None else c for c in row])
    return aoa


def find_products(aoa):
    """Find products in a CREATION / DISPATCH / EXTRACTION sheet.
    Only a Designation/Description column is required; the rest are optional."""
    hdr = -1
    for i, row in enumerate(aoa):
        low = [str(c).lower() for c in row]
        if any("designation" in c or "description" in c for c in low):
            hdr = i
            break
    if hdr < 0:
        raise ValueError("Could not find a header row with a Designation/Description column.")
    low = [str(c).strip().lower() for c in aoa[hdr]]

    def col(*names):
        for idx, c in enumerate(low):
            if any(n in c for n in names):
                return idx
        return -1

    def exact(name):
        for idx, c in enumerate(low):
            if c == name:
                return idx
        return -1

    # prefer the exact "paht" column; else the first "paht" that is NOT "paht n-1"
    paht_col = exact("paht")
    if paht_col < 0:
        paht_col = next((i for i, c in enumerate(low) if "paht" in c and "n-1" not in c), -1)

    ci = {
        "des": col("designation", "description"),
        "brand": col("brand", "marque"),
        "rayon": col("rayon", "dept"),
        "model": col("modele", "ref frs", "ref_frs", "modele fr"),
        "ean": col("ean", "barcode", "codebarre", "code barre", "code-barre", "ref fourn"),
        "color": col("couleur", "color", "colour"),
        "size": col("taille", "size"),
        "qty": col("qte", "qty", "quantit"),
        "paht": paht_col,
    }

    def g(row, idx):
        if idx < 0 or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    out = []
    for i in range(hdr + 1, len(aoa)):
        row = aoa[i]
        des = g(row, ci["des"])
        if not des:
            continue
        out.append({
            "no": len(out) + 1, "designation": des,
            "brand": g(row, ci["brand"]), "rayon": g(row, ci["rayon"]),
            "model": g(row, ci["model"]), "ean": g(row, ci["ean"]),
            "color": g(row, ci["color"]), "size": g(row, ci["size"]),
            "qty": g(row, ci["qty"]), "paht": g(row, ci["paht"]),
        })
    return out


# ---------------------------------------------------------------- building
def _creation_aoa(rows, include_enrich, include_barcode):
    head = ["FOURNISSEUR", "MARQUE", "SAISON", "ANNEE", "RAYON", "CAT / FAMILLE", "Description",
            "TAILLE", "COULEUR", "REF FAI", "REF FOURNISSEUR", "QTES", "PAHT", "PVC", "PV FAI", "EAN FAI"]
    if include_enrich:
        head.append("ENRICHED DESCRIPTION")
    if include_barcode:
        head.append("BARCODE (EAN)")
    aoa = [["CREATION FILE"], head]
    for r in rows:
        # REF FOURNISSEUR = the supplier reference (ref_frs); REF FAI = ref_n1 (usually empty);
        # the EAN/barcode goes to the EAN FAI column (client feedback 2026-07-28).
        row = [r.get("supplier", ""), r.get("brand", ""), r.get("season", ""), r.get("year", ""),
               to_rayon(r.get("dept", "")), r.get("cat_family", ""), r.get("designation", ""),
               r.get("size", ""), r.get("color", ""), r.get("ref_n1", ""), str(r.get("ref_frs", "") or ""),
               num(r.get("total_qty")), num(r.get("paht")), num(r.get("pvc")), num(r.get("pv_fai")),
               str(r.get("barcode", "") or "")]
        if include_enrich:
            row.append(r.get("enriched_data", "") or "")
        if include_barcode:
            row.append(str(r.get("barcode", "") or ""))
        aoa.append(row)
    return aoa, {12: EUR_FMT}   # PAHT column


def _dispatch_aoa(rows):
    head = ["RAYON", "REF N-1", "MODELE FR", "REF FOURN", "DESIGNATION", "COLISAGE", "NB COLIS", "QTES",
            "PAHT N-1", "PAHT", "TOTAL", "Price Benchmarking : Average Price",
            "Price Benchmarking : Highest Price", "Price Benchmarking : Lowest Price",
            "PVC N-1", "PVC", "PV FAI N-1", "PV FAI", "MAG 1", "MAG 2", "MAG 3", "MAG 4", "MAG 5"]
    aoa = [["DISPATCH FILE"], head]
    for r in rows:
        modele = str(r.get("ref_frs", "") or "").split("/")[0]
        paht, qty = num(r.get("paht")), num(r.get("total_qty"))
        total = paht * qty if (paht != "" and qty != "") else ""
        aoa.append([to_rayon(r.get("dept", "")), r.get("ref_n1", ""), modele, str(r.get("ref_frs", "") or ""),
                    r.get("designation", ""), r.get("colisage", "") or "", "", qty, "", paht, total,
                    num(r.get("price_avg")), num(r.get("price_high")), num(r.get("price_low")),
                    "", num(r.get("pvc")), "", "", "", "", "", "", ""])
    return aoa, {9: EUR_FMT, 10: EUR_FMT}   # PAHT, TOTAL


def _dispatch_benchmark_aoa(rows):
    head = ["RAYON", "REF N-1", "MODELE FR", "REF FOURN", "DESIGNATION", "COLISAGE", "NB COLIS", "QTES",
            "PAHT N-1", "PAHT", "TOTAL", "Price Benchmarking : Average Price",
            "Price Benchmarking : Highest Price", "Price Benchmarking : Lowest Price",
            "PVC N-1", "PVC", "PV FAI N-1", "PV FAI", "MAG 1", "MAG 2", "MAG 3", "MAG 4", "MAG 5",
            "Benchmarking References"]
    aoa = [["DISPATCH FILE"], head]
    for r in rows:
        paht, qty = num(r.get("paht")), num(r.get("qty"))
        total = paht * qty if (paht != "" and qty != "") else ""
        aoa.append([r.get("rayon", "") or "", "", r.get("model", "") or "", str(r.get("ean", "") or ""),
                    r.get("designation", "") or "", "", "", qty, "", paht, total,
                    num(r.get("price_avg")), num(r.get("price_high")), num(r.get("price_low")),
                    "", "", "", "", "", "", "", "", "", r.get("benchmarking_reference", "") or ""])
    return aoa, {9: EUR_FMT, 10: EUR0_FMT, 11: EUR_FMT, 12: EUR_FMT, 13: EUR_FMT}


def _build_xlsx(sheets):
    """sheets: list of (name, aoa, fmt_cols). fmt_cols: {0-based col index: number_format}.
    Number formats apply to data rows (row 3+) whose value is numeric."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, aoa, fmt_cols in sheets:
        ws = wb.create_sheet(title=name)
        for r_i, row in enumerate(aoa):
            for c_i, val in enumerate(row):
                cell = ws.cell(row=r_i + 1, column=c_i + 1, value=(val if val != "" else None))
                if r_i >= 2 and c_i in fmt_cols and isinstance(val, (int, float)):
                    cell.number_format = fmt_cols[c_i]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _aoa_to_csv(aoa):
    """Raw-number CSV (no € formatting) with a UTF-8 BOM — safe for direct import."""
    buf = io.StringIO()
    w = _csv.writer(buf)
    for row in aoa:
        w.writerow(["" if v == "" else v for v in row])
    return "﻿" + buf.getvalue()


def build_creation_dispatch(rows, include_enrich=False, include_barcode=False):
    """Return {'xlsx': bytes, 'creation_csv': str, 'dispatch_csv': str}."""
    c_aoa, c_fmt = _creation_aoa(rows, include_enrich, include_barcode)
    d_aoa, d_fmt = _dispatch_aoa(rows)
    return {
        "xlsx": _build_xlsx([("CREATION", c_aoa, c_fmt), ("DISPATCH", d_aoa, d_fmt)]),
        "creation_csv": _aoa_to_csv(c_aoa),
        "dispatch_csv": _aoa_to_csv(d_aoa),
    }


def build_dispatch_benchmark(rows):
    """Return {'xlsx': bytes} for the benchmark DISPATCH output."""
    aoa, fmt = _dispatch_benchmark_aoa(rows)
    return {"xlsx": _build_xlsx([("DISPATCH", aoa, fmt)])}
